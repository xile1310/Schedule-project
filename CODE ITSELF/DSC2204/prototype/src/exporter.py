"""
Segment 5 — Output

Produces:
* `timetable.json`  — machine-readable schedule (the canonical artefact)
* `violations.json` — constraint engine report
* `schedule_output.html` — single-file static schedule snapshot with stakeholder views

The HTML is intentionally self-contained (no external assets) so it
can be opened from disk, emailed, or embedded in a share drive.
"""
from __future__ import annotations
import json, html
from pathlib import Path
from typing import Dict, List
from collections import defaultdict

from .models import Timetable, Universe
from .constraint_engine import ViolationReport


# ---------------------------------------------------------------------------

def write_json(timetable: Timetable, report: ViolationReport, out_dir: Path):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "timetable.json").write_text(timetable.to_json(), encoding="utf-8")
    (out_dir / "violations.json").write_text(json.dumps(report.as_dict(), indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# HTML dashboard
# ---------------------------------------------------------------------------

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
SLOT_MIN = 30
DAY_START_HOUR = 8


def write_dashboard(timetable: Timetable, report: ViolationReport,
                    universe: Universe, out_path: Path):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "assignments": [a.__dict__ for a in timetable.assignments],
        "violations": report.as_dict(),
        "metadata": timetable.metadata,
        "rooms": [{"id": r.id, "name": r.name, "capacity": r.capacity,
                   "type": r.room_type.value} for r in universe.rooms],
        "tutors": [{"id": t.id, "name": t.name} for t in universe.tutors],
        "groups": [{"id": g.id, "label": g.label, "size": g.size,
                    "course": g.course_code} for g in universe.groups],
        "courses": [{"code": c.code, "year": c.year, "programme": c.programme}
                    for c in universe.courses],
        "days": DAYS,
        "slot_min": SLOT_MIN,
        "day_start_hour": DAY_START_HOUR,
        "n_slots": universe.slot_count,
        "weeks": universe.calendar.teaching_weeks,
        "week_dates": universe.calendar.week_dates,
    }
    out_path.write_text(_render_html(payload), encoding="utf-8")


def _render_html(payload: Dict) -> str:
    pj = json.dumps(payload)
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>DSC Timetable Dashboard</title>
<style>
  :root {{
    --bg: #0f1419;
    --panel: #1b232c;
    --muted: #8aa0b4;
    --text: #e6edf3;
    --accent: #ff5046;
    --hard: #ff5046;
    --soft: #ffb84d;
    --ok: #5fd068;
    --grid: #29333d;
    --f2f: #4cb5ff;
    --online: #b86bff;
    --highlight: #ffe066;
  }}
  * {{ box-sizing: border-box; }}
  body {{ margin:0; font:14px/1.4 -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
         background: var(--bg); color: var(--text); }}
  header {{ padding: 14px 20px; background: var(--panel); display:flex; align-items:baseline; gap:18px;
            border-bottom: 1px solid var(--grid); flex-wrap: wrap;}}
  header h1 {{ font-size: 17px; margin:0; }}
  header .pill {{ background: var(--bg); padding:4px 10px; border-radius:99px; font-size:12px;
                  color: var(--muted); border:1px solid var(--grid);}}
  header .pill.bad {{ color: var(--hard); border-color:var(--hard); }}
  header .pill.warn {{ color: var(--soft); border-color: var(--soft); }}
  header .pill.ok {{ color: var(--ok); border-color: var(--ok); }}
  main {{ display:grid; grid-template-columns: 250px 1fr; min-height: calc(100vh - 60px); }}
  aside {{ background: var(--panel); padding: 14px; border-right:1px solid var(--grid); }}
  aside h3 {{ font-size:11px; color: var(--muted); margin:14px 0 6px; letter-spacing:.08em;
              text-transform: uppercase;}}
  aside select, aside button {{ width:100%; padding:8px 10px; background: var(--bg);
                                border:1px solid var(--grid); color:var(--text); border-radius:6px;
                                margin-bottom:6px; font:inherit;}}
  aside button:hover {{ background: var(--grid); cursor:pointer; }}
  section {{ padding: 16px 20px; overflow-x:auto; }}
  .grid {{ display:grid; grid-template-columns: 60px repeat(5, 1fr); gap:1px; background:var(--grid);
           border:1px solid var(--grid); position:relative; min-width: 760px; }}
  .grid > div {{ background: var(--bg); padding: 4px 6px; min-height: 30px; font-size:12px; }}
  .grid .slot.has-event {{ position: relative; z-index: 5; overflow: visible; }}
  .grid .head {{ background: var(--panel); font-weight:600; text-align:center; padding:8px 0; }}
  .grid .timecell {{ color: var(--muted); font-size:11px; text-align:right; padding-right:8px; padding-top:2px;}}
  .grid .slot {{ min-height: 24px; }}
  .event {{ position:absolute; left:2px; right:2px; top:0; padding:6px 8px; border-radius:6px;
            color:#0e1117; font-size: 11px; line-height: 1.25; cursor:pointer;
            box-shadow: 0 1px 3px rgba(0,0,0,.35); overflow:hidden; }}
  .event.f2f {{ background: var(--f2f); }}
  .event.online_sync, .event.online_async {{ background: var(--online); color:#fff;}}
  .event .when {{ font-weight: 600; }}
  .event:hover {{ outline: 2px solid var(--highlight); }}
  .legend {{ display:flex; gap:10px; align-items:center; margin: 10px 0 6px; flex-wrap:wrap;}}
  .legend .dot {{ display:inline-block; width:10px; height:10px; border-radius:99px; margin-right:5px;}}
  .vio {{ background:#211a1a; border-left:3px solid var(--hard); padding:8px 10px; margin:6px 0;
          font-size: 12px; border-radius:0 6px 6px 0;}}
  .vio.soft {{ border-color: var(--soft); background:#221d11;}}
  .vio code {{ background: rgba(255,255,255,.06); padding:1px 5px; border-radius:3px;}}
  .empty {{ color:var(--muted); font-style: italic; padding: 14px;}}
  .tabs {{ display:flex; gap:4px; margin-bottom:10px; }}
  .tabs button {{ background: var(--panel); padding:6px 12px; border-radius:6px;
                  border:1px solid var(--grid); color:var(--muted); cursor:pointer; font:inherit;}}
  .tabs button.on {{ background: var(--accent); color:#fff; border-color: var(--accent);}}
  details summary {{ cursor:pointer; padding:4px 0; }}
</style></head>
<body>
<header>
  <h1>DSC Timetable Dashboard <span style="color:var(--muted); font-weight: normal;">— DSC2204 ITP prototype</span></h1>
  <span class="pill" id="solver"></span>
  <span class="pill" id="status"></span>
  <span class="pill" id="hardp"></span>
  <span class="pill" id="softp"></span>
</header>
<main>
  <aside>
    <h3>View</h3>
    <select id="viewMode">
      <option value="programme">By programme</option>
      <option value="year">By year</option>
      <option value="student">By student</option>
      <option value="tutor">By tutor</option>
      <option value="room">By room</option>
      <option value="group">By group</option>
    </select>
    <select id="filter"></select>
    <h3>Week</h3>
    <select id="weekFilter"><option value="all">All weeks</option></select>
    <h3>Legend</h3>
    <div class="legend"><span class="dot" style="background:var(--f2f)"></span>Face-to-face</div>
    <div class="legend"><span class="dot" style="background:var(--online)"></span>Online</div>
    <h3>Violations</h3>
    <div id="vios"></div>
  </aside>
  <section>
    <div id="title" style="font-size:15px;margin-bottom:8px;color:var(--muted);"></div>
    <div id="grid"></div>
  </section>
</main>
<script>
const DATA = {pj};

const $ = id => document.getElementById(id);
const sl2hhmm = s => {{
  const totalMin = DATA.day_start_hour*60 + s*DATA.slot_min;
  return String(Math.floor(totalMin/60)).padStart(2,'0')+':'+String(totalMin%60).padStart(2,'0');
}};

function init() {{
  $('solver').textContent = 'Solver: ' + (DATA.metadata.solver || 'n/a');
  const status = DATA.violations.summary;
  $('status').textContent = status.feasible ? 'Feasible' : 'Infeasible';
  $('status').className = 'pill ' + (status.feasible ? 'ok' : 'bad');
  $('hardp').textContent = status.hard_count + ' hard';
  $('hardp').className = 'pill ' + (status.hard_count ? 'bad' : 'ok');
  $('softp').textContent = status.soft_count + ' soft (score ' + status.soft_score + ')';
  $('softp').className = 'pill ' + (status.soft_count ? 'warn' : 'ok');

  // weeks — labelled with the actual Monday-of-week date
  const wf = $('weekFilter');
  const fmtShort = iso => {{
    const d = new Date(iso);
    const m = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
    return d.getDate() + ' ' + m[d.getMonth()];
  }};
  DATA.weeks.forEach(w => {{
    const o = document.createElement('option');
    o.value = w;
    const dt = (DATA.week_dates || {{}})[w];
    o.textContent = 'Week ' + w + (dt ? ' (' + fmtShort(dt) + ')' : '');
    wf.appendChild(o);
  }});

  $('viewMode').onchange = renderControls;
  $('filter').onchange = render;
  $('weekFilter').onchange = render;

  renderViolations();
  renderControls();
}}

const COURSE_YEAR = Object.fromEntries(DATA.courses.map(c => [c.code, c.year]));
const COURSE_PROG = Object.fromEntries(DATA.courses.map(c => [c.code, c.programme || c.code.slice(0,3)]));

function uniqueValues(mode) {{
  if (mode === 'programme') return [...new Set(DATA.assignments.map(a => COURSE_PROG[a.course_code] || a.course_code.slice(0,3)))].sort();
  if (mode === 'year') {{
    const ys = [...new Set(DATA.assignments.map(a => COURSE_YEAR[a.course_code] || '?'))].sort();
    return ys.map(y => 'Year ' + y);
  }}
  if (mode === 'student') {{
    const out = new Set();
    DATA.assignments.forEach(a => {{
      const y = COURSE_YEAR[a.course_code] || '?';
      const label = a.group_id.split('/').pop();
      if (label === 'All') return;
      const m = label.match(/^[A-Za-z]+(\d+)$/);
      if (m) out.add('Year ' + y + ' \u2014 sub-group ' + m[1]);
    }});
    return [...out].sort();
  }}
  if (mode === 'tutor') return [...new Set(DATA.assignments.map(a => a.tutor_name))].sort();
  if (mode === 'room') return [...new Set(DATA.assignments.map(a => a.room_id))].sort();
  if (mode === 'group') return [...new Set(DATA.assignments.map(a => a.group_id))].sort();
  return [];
}}

function renderControls() {{
  const mode = $('viewMode').value;
  const f = $('filter');
  f.innerHTML = '<option value="__all__">All</option>';
  if (mode === 'group') {{
    // Group by module code so the dropdown is browsable.
    const byModule = {{}};
    uniqueValues(mode).forEach(v => {{
      const mod = v.split('/')[0];
      (byModule[mod] = byModule[mod] || []).push(v);
    }});
    Object.keys(byModule).sort().forEach(mod => {{
      const og = document.createElement('optgroup');
      og.label = mod;
      byModule[mod].forEach(v => {{
        const o = document.createElement('option');
        o.value = v;
        o.textContent = v.split('/').slice(1).join('/');
        og.appendChild(o);
      }});
      f.appendChild(og);
    }});
  }} else {{
    uniqueValues(mode).forEach(v => {{
      const o = document.createElement('option'); o.value=v; o.textContent=v; f.appendChild(o);
    }});
  }}
  // Auto-select the first real option so the calendar immediately filters
  if (f.options.length > 1) f.selectedIndex = 1;
  render();
}}

function filterAssignments() {{
  const mode = $('viewMode').value;
  const v = $('filter').value;
  const w = $('weekFilter').value;
  return DATA.assignments.filter(a => {{
    if (v !== '__all__') {{
      if (mode==='programme' && (COURSE_PROG[a.course_code] || a.course_code.slice(0,3)) !== v) return false;
      if (mode==='year'      && ('Year ' + (COURSE_YEAR[a.course_code]||'?')) !== v) return false;
      if (mode==='student') {{
        const m = v.match(/^Year (\S+) \u2014 sub-group (\d+)$/);
        if (m) {{
          const studentYear = String(COURSE_YEAR[a.course_code] || '?');
          if (studentYear !== m[1]) return false;
          const label = a.group_id.split('/').pop();
          if (label !== 'All') {{
            const sm = label.match(/^[A-Za-z]+(\d+)$/);
            if (!sm || sm[1] !== m[2]) return false;
          }}
        }}
      }}
      if (mode==='tutor'     && a.tutor_name !== v) return false;
      if (mode==='room'      && a.room_id !== v) return false;
      if (mode==='group'     && a.group_id !== v) return false;
    }}
    if (w !== 'all' && !a.weeks.includes(parseInt(w))) return false;
    return true;
  }});
}}

function render() {{
  const items = filterAssignments();
  const grid = $('grid');
  grid.innerHTML = '';
  grid.className = 'grid';
  // header row — include the date when a single week is selected
  grid.appendChild(cell('','head'));
  const wkSel = $('weekFilter').value;
  const wkMon = (DATA.week_dates || {{}})[wkSel];
  const dayDates = {{}};
  if (wkMon) {{
    const base = new Date(wkMon);
    DATA.days.forEach((d, i) => {{
      const dt = new Date(base); dt.setDate(base.getDate() + i);
      const m = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
      dayDates[d] = dt.getDate() + ' ' + m[dt.getMonth()];
    }});
  }}
  DATA.days.forEach(d => {{
    const label = wkMon ? (d + ' · ' + dayDates[d]) : d;
    grid.appendChild(cell(label, 'head'));
  }});
  // body — one row per slot
  for (let s = 0; s < DATA.n_slots; s++) {{
    grid.appendChild(cell(sl2hhmm(s),'timecell'));
    DATA.days.forEach(d => grid.appendChild(cell('','slot')));
  }}
  // overlay events (positioned absolutely inside the cell)
  // Find each slot div by its index: header (1+5) + row*(1+5) + (1+dayIdx)
  const colCount = 6;

  // Pre-group items by (day, start_index) so we can split overlapping events
  // into side-by-side columns (like a real calendar app).
  const overlapKey = a => a.day + '|' + a.start_index;
  const overlapGroups = {{}};
  items.forEach(a => {{
    const k = overlapKey(a);
    (overlapGroups[k] = overlapGroups[k] || []).push(a);
  }});

  items.forEach(a => {{
    const dayIdx = DATA.days.indexOf(a.day);
    if (dayIdx < 0) return;
    const cellIdx = colCount + a.start_index*colCount + (1+dayIdx);
    const c = grid.children[cellIdx];
    if (!c) return;
    c.classList.add('has-event');
    const ev = document.createElement('div');
    ev.className = 'event ' + a.delivery_mode;
    ev.style.height = (a.duration_slots * 30 - 4) + 'px';

    // Side-by-side layout when multiple events share the same start slot
    const group = overlapGroups[overlapKey(a)];
    const nCols = group.length;
    const colIdx = group.indexOf(a);
    if (nCols > 1) {{
      const pct = 100 / nCols;
      ev.style.left  = `calc(${{colIdx * pct}}% + 2px)`;
      ev.style.right = `calc(${{(nCols - colIdx - 1) * pct}}% + 2px)`;
      ev.style.top   = '0';
    }}

    const notesHtml = (a.notes && a.notes.trim())
      ? `<div style="opacity:.75;font-style:italic;margin-top:2px">${{a.notes.trim()}}</div>` : '';
    ev.innerHTML = `<div class="when">${{a.start_label.slice(0,2)}}:${{a.start_label.slice(2)}} – ${{a.end_label.slice(0,2)}}:${{a.end_label.slice(2)}}</div>
      <div><b>${{a.course_code}}</b> ${{a.activity_type}}</div>
      <div>${{a.group_id.split('/').pop()}} · ${{a.tutor_name}}</div>
      <div style="opacity:.85">${{a.room_id}}</div>
      ${{notesHtml}}`;
    ev.title = JSON.stringify(a, null, 2);
    c.appendChild(ev);
  }});
  $('title').textContent = `${{items.length}} sessions shown`;
}}

function cell(text, cls) {{
  const d = document.createElement('div');
  d.className = cls; d.textContent = text;
  return d;
}}

function renderViolations() {{
  const c = $('vios');
  const all = DATA.violations.hard.concat(DATA.violations.soft);
  if (!all.length) {{ c.innerHTML = '<div class="empty">No violations ✅</div>'; return; }}
  c.innerHTML = '';
  all.slice(0, 60).forEach(v => {{
    const div = document.createElement('div');
    div.className = 'vio ' + v.severity;
    div.innerHTML = `<code>${{v.code}}</code> ${{v.message}}`;
    c.appendChild(div);
  }});
  if (all.length > 60) {{
    const more = document.createElement('div');
    more.className='empty';
    more.textContent = `(+${{all.length-60}} more — see violations.json)`;
    c.appendChild(more);
  }}
}}

init();
</script></body></html>
"""


# ---------------------------------------------------------------------------
# Excel output — the solved timetable as a spreadsheet, plus a "who teaches
# what" sheet.  Also a helper to write the solver's chosen staff back into
# the Module tab's Staff columns of the input workbook.
# ---------------------------------------------------------------------------

def write_timetable_xlsx(timetable, universe, out_path):
    """Write the solved timetable to an .xlsx with two sheets:
       'Timetable'        — one row per scheduled session (day/time/room/tutor)
       'Staff Assignments'— one row per module-activity with the assigned staff."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    hdr_font = Font(bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF217346")
    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    year_of = {c.code: c.year for c in universe.courses}
    day_order = {d: i for i, d in enumerate(universe.days)}
    wkstr = lambda ws: ",".join(str(w) for w in ws)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: full timetable ----
    ws = wb.active
    ws.title = "Timetable"
    cols = ["Module", "Yr", "Activity", "Group", "Delivery Mode", "Day",
            "Start", "End", "Weeks", "Room", "Tutor ID", "Tutor", "Remarks"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    rows = sorted(timetable.assignments,
                  key=lambda a: (day_order.get(a.day, 9), a.start_index,
                                 a.course_code, a.group_id))
    for i, a in enumerate(rows, start=2):
        vals = [a.course_code, year_of.get(a.course_code, ""), a.activity_type,
                a.group_id.split("/")[-1], a.delivery_mode, a.day,
                a.start_label, a.end_label, wkstr(a.weeks), a.room_id,
                a.tutor_id,
                (a.tutor_name + (" + " + ", ".join(a.co_tutor_names) if a.co_tutor_names else "")),
                getattr(a, "notes", "")]
        for j, val in enumerate(vals, 1):
            cell = ws.cell(i, j, val); cell.border = border
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    for j, w in enumerate([10, 4, 11, 7, 14, 5, 7, 7, 24, 24, 10, 22, 40], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---- Sheet 2: who teaches what ----
    ws2 = wb.create_sheet("Staff Assignments")
    c2 = ["Module", "Activity", "Delivery Mode", "Weeks", "Assigned Staff (ID)"]
    for j, c in enumerate(c2, 1):
        cell = ws2.cell(1, j, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    agg = defaultdict(lambda: {"weeks": set(), "staff": set(), "mode": ""})
    for a in timetable.assignments:
        k = (a.course_code, a.activity_type)
        agg[k]["weeks"].update(a.weeks)
        agg[k]["staff"].add(f"{a.tutor_name} ({a.tutor_id})")
        for _cid, _cn in zip(a.co_tutor_ids, a.co_tutor_names):
            agg[k]["staff"].add(f"{_cn} ({_cid})")
        agg[k]["mode"] = a.delivery_mode
    for i, (k, d) in enumerate(sorted(agg.items()), start=2):
        vals = [k[0], k[1], d["mode"],
                ",".join(str(w) for w in sorted(d["weeks"])),
                " / ".join(sorted(d["staff"]))]
        for j, val in enumerate(vals, 1):
            cell = ws2.cell(i, j, val); cell.border = border
    ws2.freeze_panes = "A2"
    for j, w in enumerate([10, 12, 16, 26, 46], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_simple_results(timetable, universe, out_path):
    """Single-sheet, human-friendly summary of the solved timetable.

    Columns: Course, Activity, Tutor, Room, Day, Start, End, Duration (h),
    Weeks, Size.  Auto-regenerated by run.py every time the solver runs,
    so the file always reflects the latest schedule.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    hdr_font = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF1F3864")
    body_font = Font(name=FONT, size=10)
    alt_fill = PatternFill("solid", fgColor="FFF2F2F2")
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    activity_colors = {
        "Lecture": "FFFFD966", "Tutorial": "FFA9D08E", "Laboratory": "FF9BC2E6",
        "Workshop": "FFF4B084", "Quiz": "FFFFA9A9", "Lectorial": "FFC9A6E0",
        "Seminar": "FFD0CECE", "Other": "FFBFBFBF",
    }

    day_order = {d: i for i, d in enumerate(universe.days)}
    rows = sorted(timetable.assignments,
                  key=lambda a: (day_order.get(a.day, 9), a.start_index,
                                 a.course_code, a.group_id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    ws.sheet_view.showGridLines = False

    cols = ["Course", "Activity", "Tutor", "Room", "Day",
            "Start", "End", "Duration (h)", "Weeks", "Size"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, a in enumerate(rows, start=2):
        tutor_disp = a.tutor_name + (
            " + " + ", ".join(a.co_tutor_names) if a.co_tutor_names else "")
        vals = [
            a.course_code,
            a.activity_type,
            tutor_disp,
            a.room_name,
            a.day,
            a.start_label,
            a.end_label,
            round(a.duration_slots * 0.5, 1),
            ",".join(str(w) for w in a.weeks),
            a.size,
        ]
        for j, val in enumerate(vals, 1):
            cell = ws.cell(i, j, val)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if j in (5, 6, 7, 8, 10) else "left")
            if (i % 2) == 1:
                cell.fill = alt_fill
        # colour the Activity cell by activity type
        ws.cell(i, 2).fill = PatternFill("solid",
            fgColor=activity_colors.get(a.activity_type, "FFBFBFBF"))

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"

    widths = [10, 12, 28, 26, 6, 7, 7, 12, 22, 6]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Template 2 — fill the SIT planning template's "Timetable" sheet with the
# solver's output, preserving every other reference sheet in the workbook.
# ---------------------------------------------------------------------------

_CLASS_TYPE_TO_CODE = {
    "Lecture":   "LEC",
    "Tutorial":  "TUT",
    "Laboratory":"LAB",
    "Workshop":  "WOR",
    "Quiz":      "QUZ",
    "Seminar":   "SEM",
    "Lectorial": "LET",
    "Practicum": "PRA",
    "Assignment":"ASS",
    "Clinical":  "CLN",
    "Discussion":"DIS",
    "Field_Studies":   "FLD",
    "Fieldwork":       "FLW",
    "Independent_Study":"IND",
    "Preparatory_Work": "PRE",
    "Projects":  "PRJ",
    "Research":  "RSC",
    "Self_Study":"SES",
    "Supervision":"SUP",
    "Other":     "BKG",
}

def _sector_for_room(room, default_campus: str) -> str:
    """Derive Sector from the assigned room's own zone column.

    The Rooms tab stores each room's zone (e.g. 'E3' for Punggol building E3,
    'DV' for Dover). We map those to PUNGGOL / DOVER. Nothing is hardcoded by
    module-code prefix; whichever room the solver picked decides the sector.

    For the synthetic VIRTUAL room (online classes), we fall back to
    `default_campus` — typically the majority campus of the worksheet's
    physical rooms, inferred at write time.
    """
    if room is None:
        return default_campus
    zone = (room.zone or "").strip().upper()
    if zone.startswith("E") and zone[1:2].isdigit():   # E1..E9 = Punggol
        return "PUNGGOL"
    if zone.startswith("DV"):
        return "DOVER"
    if zone in ("ONLINE", ""):
        return default_campus
    # Anything else: keep the zone text verbatim so the planner can see it
    return zone


def _infer_default_campus(universe) -> str:
    """Look at the physical rooms in the workbook and pick the dominant
    campus. Used as the sector for online classes (no real room)."""
    counts = {"PUNGGOL": 0, "DOVER": 0}
    for r in getattr(universe, "rooms", []) or []:
        if getattr(r, "is_virtual", False):
            continue
        zone = (r.zone or "").upper()
        if zone.startswith("E") and zone[1:2].isdigit():
            counts["PUNGGOL"] += 1
        elif zone.startswith("DV"):
            counts["DOVER"] += 1
    return max(counts, key=counts.get) if any(counts.values()) else "PUNGGOL"


def write_template2(timetable, universe, template_in, out_path,
                    term=2520, cluster_tag="ENG-UGRD-PU"):
    """Populate template_in's Timetable sheet from timetable.assignments and
    write to out_path.  Every other sheet in the workbook (Course Code,
    Location, Staff, Group, Zone, Time, Day, Class Type, Template,
    StaffGroup, ...) is preserved untouched."""
    import openpyxl, warnings, zipfile, os
    from collections import defaultdict
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

    warnings.simplefilter("ignore")

    HEADERS = ["Module", "Class Type", "Template", "Group", "Day", "Start",
               "End", "Class Size", "Sector", "RoomGrouping", "Room1",
               "Room2", "StaffGrouping", "Staff1", "Staff2", "Tri Week",
               "Recording Mode", "Remark", "FMTS Tri Start Week",
               "Activity Hostkey", "SIS Module Code", "Term", "Activity Type",
               "Duration", "Staff Suitability ID", "SIS Staff ID",
               "SIS Staff ID", "Zone Hoskey", "Location Suitability ID",
               "Location Hostkey", "Location Hostkey"]

    def _new_blank_book():
        """Build a fresh Output workbook with just the Timetable sheet +
        headers, used when the on-disk file is missing or corrupt."""
        nwb = openpyxl.Workbook()
        nws = nwb.active
        nws.title = "Timetable"
        hdr_fill = PatternFill("solid", fgColor="FF1F3864")
        hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
        for j, h in enumerate(HEADERS, 1):
            cell = nws.cell(1, j, h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        return nwb

    try:
        wb = openpyxl.load_workbook(template_in)
        if "Timetable" not in wb.sheetnames:
            raise ValueError("no Timetable sheet")
    except (zipfile.BadZipFile, FileNotFoundError, KeyError, ValueError, OSError):
        # Existing Output.xlsx is missing or corrupt — build a fresh one.
        wb = _new_blank_book()
    ws = wb["Timetable"]

    # Make sure the header row exists and matches the expected schema.
    if ws.cell(1, 1).value is None:
        for j, h in enumerate(HEADERS, 1):
            ws.cell(1, j, h)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    template_seq = defaultdict(int)
    day_order = {d: i for i, d in enumerate(universe.days)}
    rows = sorted(timetable.assignments,
                  key=lambda a: (a.course_code, a.activity_type,
                                 a.group_id, day_order.get(a.day, 9),
                                 a.start_index))

    # Look up each room object by id for zone-based sector lookup.
    room_by_id = {r.id: r for r in (universe.rooms or [])}
    default_campus = _infer_default_campus(universe)

    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name="Arial", size=10)

    for a in rows:
        seq_key = (a.course_code, a.activity_type, a.group_id)
        template_seq[seq_key] += 1
        tnum = template_seq[seq_key]
        group_label = a.group_id.split("/", 1)[-1] if "/" in a.group_id else a.group_id
        weeks_str = ",".join(str(w) for w in a.weeks)
        first_week = min(a.weeks) if a.weeks else 1
        sector = _sector_for_room(room_by_id.get(a.room_id), default_campus)
        sis_code = f"{a.course_code}-{term}-{cluster_tag}"
        act_code = _CLASS_TYPE_TO_CODE.get(a.activity_type, "BKG")
        activity_hk = f"{sis_code}-{act_code}/{group_label}"
        duration_periods = int(round(a.duration_slots * 30 / 20))
        rec_mode = "A0" if str(a.delivery_mode).lower() == "f2f" else ""
        room_hostkey = a.room_id if a.room_id else ""
        room2_hostkey = a.room2_id if getattr(a, "room2_id", "") else ""
        shared = getattr(a, "shared_cohorts", [])
        if shared:
            remark = "Combined: " + " + ".join(shared)
        else:
            remark = a.notes or ""
        co_tutor_name = a.co_tutor_names[0] if a.co_tutor_names else ""
        co_tutor_id = a.co_tutor_ids[0] if a.co_tutor_ids else ""

        ws.append([
            a.course_code, a.activity_type, tnum, group_label, a.day,
            a.start_label, a.end_label, a.size, sector, "",
            room_hostkey, room2_hostkey, "", a.tutor_name, co_tutor_name,
            weeks_str, rec_mode, remark, first_week, activity_hk,
            sis_code, term, act_code, duration_periods, "",
            a.tutor_id, co_tutor_id, sector, "", room_hostkey, room2_hostkey,
        ])

    for r in range(2, ws.max_row + 1):
        for c in range(1, 32):
            cell = ws.cell(r, c)
            cell.font = body_font
            cell.border = border

    ws.freeze_panes = "A2"
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_back_staff(timetable, worksheet_path, sheet="Module"):
    """Best-effort write-back of the Staff columns into the source workbook's
    Module tab.  Returns True if any cells were updated, False otherwise.
    Safe to skip on errors — write_back_staff is convenience, not core."""
    import openpyxl, warnings
    from collections import defaultdict

    warnings.simplefilter("ignore")
    try:
        wb = openpyxl.load_workbook(worksheet_path)
    except Exception:
        return False
    if sheet not in wb.sheetnames:
        return False
    ws = wb[sheet]

    asgs = defaultdict(list)
    for a in timetable.assignments:
        key = (a.course_code, a.activity_type, ",".join(str(w) for w in a.weeks))
        asgs[key].append((a.tutor_id, a.tutor_name))

    hdr_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            hdr_map[str(v).strip().lower()] = c

    cc = hdr_map.get("module code")
    ca = hdr_map.get("activity")
    cw = hdr_map.get("teaching weeks")
    if not (cc and ca and cw):
        return False

    import re as _re
    touched = False
    cur_code = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, cc).value
        if v:
            cur_code = str(v).strip()
        act = ws.cell(r, ca).value
        wks = ws.cell(r, cw).value
        if not (cur_code and act):
            continue
        key = (cur_code, str(act).strip(), str(wks).strip() if wks else "")
        if key not in asgs:
            continue
        pairs = []
        for hdr, ci in hdr_map.items():
            m = _re.match(r"^staff\s+(\d+)$", hdr)
            if m:
                n = int(m.group(1))
                pairs.append((n, ci, hdr_map.get(f"staff id {n}")))
        pairs.sort(key=lambda x: x[0])
        for i, (tid, tname) in enumerate(asgs[key]):
            if i >= len(pairs):
                break
            _, name_ci, id_ci = pairs[i]
            if name_ci:
                ws.cell(r, name_ci).value = tname
                touched = True
            if id_ci:
                ws.cell(r, id_ci).value = tid
                touched = True

    if touched:
        try:
            wb.save(worksheet_path)
        except Exception:
            return False
    return touched


# --- end of exporter.py ---
