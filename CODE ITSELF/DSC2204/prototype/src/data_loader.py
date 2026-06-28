"""
Loads the DSC scope from the supplied Excel workbooks into the
`Universe` object defined in `models.py`.

The team's brief is to start with the DSC programme.  The Module sheet
in `Worksheet in ITP Project Requirements.xlsx` gives us:
    DSC1001  (Yr1, 80 students) — Lecture (online sync) + Tutorial (f2f)
    DSC2302  (Yr2, 70 students) — Lecture (online sync) + Lab (f2f)
    MET2602  (Yr2, 70 students) — Lecture (online sync) + Lab (f2f)
    DSC3002B (Yr3, 67 students) — Workshop (f2f)

We also load the Dover-campus rooms from the Location sheet because
DSC sits at the Dover (DV) campus, and we add a synthetic VIRTUAL room
for online-synchronous classes.

The loader is deliberately defensive — every assumption about the
spreadsheet layout is guarded so an unfamiliar reader can debug.
"""
from __future__ import annotations
import re
import warnings
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Tuple

import openpyxl

from .models import (
    Activity, ActivityType, Calendar, Course, DeliveryMode, Group,
    Room, RoomType, TimeSlot, Tutor, Universe,
)

warnings.filterwarnings("ignore")

# ----- constants tuneable from the spreadsheets --------------------------

DEFAULT_TUTORIAL_CAP = 25
DEFAULT_LAB_CAP = 20
DEFAULT_WORKSHOP_CAP = 70
DEFAULT_LECTURE_DURATION = 4    # 4 × 30 min = 2 h
DEFAULT_TUTORIAL_DURATION = 4
DEFAULT_LAB_DURATION = 6        # 3 h
DEFAULT_WORKSHOP_DURATION = 6
DEFAULT_OTHER_DURATION = 4
DEFAULT_CLASS_SIZE = 40   # used when a Module row leaves Class Size blank

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
DAY_START_HOUR = 8
DAY_END_HOUR = 22
SLOT_MIN = 30

VIRTUAL_ROOM_ID = "VIRTUAL"

# First Monday of teaching Week 1 — used to map (week, day) → real calendar date.
SEMESTER_START_DATE = date(2025, 9, 8)   # Mon 8 Sep 2025


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _parse_weeks(raw) -> List[int]:
    if raw is None:
        return list(range(1, 14))
    if isinstance(raw, int):
        return [raw]
    weeks: list[int] = []
    for tok in re.split(r"[,;]", str(raw)):
        m = re.match(r"\s*(\d+)", tok)
        if m:
            weeks.append(int(m.group(1)))
    return sorted(set(weeks)) or list(range(1, 14))


def _split_into_n_groups(total_size: int, n: int, prefix: str):
    """Split a class into exactly N sub-groups of near-equal size, labelled
    prefix1, prefix2, ...  N comes from the Module tab's 'Subgroups' column
    (admin-provided per Prof Yang)."""
    n = max(1, int(n))
    base = total_size // n
    rem = total_size - base * n
    return [(f"{prefix}{i + 1}", base + (1 if i < rem else 0)) for i in range(n)]


def _split_into_groups(total_size: int, group_cap: int, prefix: str) -> List[Tuple[str, int]]:
    n_groups = max(1, -(-total_size // group_cap))
    base = total_size // n_groups
    rem = total_size - base * n_groups
    return [(f"{prefix}{i + 1}", base + (1 if i < rem else 0)) for i in range(n_groups)]


def _build_time_slots() -> dict:
    slots: dict[str, list[TimeSlot]] = {}
    n_per_day = (DAY_END_HOUR - DAY_START_HOUR) * 60 // SLOT_MIN
    for d in DAYS:
        day_slots = []
        for i in range(n_per_day):
            start = DAY_START_HOUR * 60 + i * SLOT_MIN
            day_slots.append(TimeSlot(day=d, index=i, start_min=start, end_min=start + SLOT_MIN))
        slots[d] = day_slots
    return slots


def _classify_room(name: str, suitabilities: str) -> RoomType:
    s = (suitabilities or "").lower()
    n = (name or "").lower()
    if "computer lab" in s:
        return RoomType.COMPUTER_LAB
    if "laboratory" in s or "lab" in n:
        return RoomType.LABORATORY
    if "lecture theatre" in s or "lt-" in n:
        return RoomType.LECTURE_THEATRE
    if "seminar" in s or "sr-" in n:
        return RoomType.SEMINAR_ROOM
    return RoomType.OTHER


# ---------------------------------------------------------------------------
# Parse free-text Remarks into scheduling pins
# ---------------------------------------------------------------------------



def _default_duration(a: ActivityType) -> int:
    return {
        ActivityType.LECTURE: DEFAULT_LECTURE_DURATION,
        ActivityType.TUTORIAL: DEFAULT_TUTORIAL_DURATION,
        ActivityType.LAB: DEFAULT_LAB_DURATION,
        ActivityType.WORKSHOP: DEFAULT_WORKSHOP_DURATION,
    }.get(a, DEFAULT_OTHER_DURATION)


def _group_cap(a: ActivityType) -> int:
    return {
        ActivityType.TUTORIAL: DEFAULT_TUTORIAL_CAP,
        ActivityType.LAB: DEFAULT_LAB_CAP,
        ActivityType.WORKSHOP: DEFAULT_WORKSHOP_CAP,
    }.get(a, 30)


def _group_prefix(a: ActivityType) -> str:
    return {
        ActivityType.TUTORIAL: "T",
        ActivityType.LAB: "L",
        ActivityType.WORKSHOP: "W",
        ActivityType.SEMINAR: "S",
    }.get(a, "G")



# ---------------------------------------------------------------------------
# Standard Period Block — canonical SIT engineering-cluster start slots
# (extracted from "Standard Period Block" sheet, ENG section)
#
# slot_index = (hour - DAY_START_HOUR) * 2 + (1 if minute>=30 else 0)
# DAY_START_HOUR = 8  → slot 2 = 09:00, slot 6 = 11:00, slot 12 = 14:00, etc.
# ---------------------------------------------------------------------------

# 2-hour blocks (Lectures, Tutorials, "Others") — ENG cluster row of the
# Standard Period Block sheet.
#   Period 1  09:00 – 11:00   (slot 2)
#   Period 2  12:00 – 14:00   (slot 8)  ← lunch window; allowed if a 1h gap exists elsewhere in 11:00–14:00
#   Period 3  14:00 – 16:00   (slot 12)
#   Period 4  16:00 – 18:00   (slot 16)
PERIOD_STARTS_2H = [2, 8, 12, 16]

# 3-hour blocks (Labs and Workshops) — ENG "Labs" column.
# Lunch is 12:00–13:00 between Period AM and Period PM.
#   Period AM 09:00 – 12:00   (slot 2)
#   Lunch     12:00 – 13:00   (skipped)
#   Period PM 13:00 – 16:00   (slot 10)
#   Period EV 16:00 – 19:00   (slot 16)
PERIOD_STARTS_3H = [2, 10, 16]


def canonical_starts(duration_slots: int) -> List[int]:
    """Return the SIT-canonical start slots for an activity of `duration_slots`.

    A duration of 4 (= 2h) returns the lecture/tutorial period starts;
    a duration of 6 (= 3h) returns the lab/workshop period starts.
    Anything else falls back to "no restriction" (returns []) so the solver
    keeps full freedom for non-standard durations.
    """
    if duration_slots == 4:
        return list(PERIOD_STARTS_2H)
    if duration_slots == 6:
        return list(PERIOD_STARTS_3H)
    return []

# ---------------------------------------------------------------------------
# Public loader
# ---------------------------------------------------------------------------

def load_dsc_universe(xlsx_modules: str, xlsx_resources: str) -> Universe:
    """Build a Universe restricted to the DSC programme."""
    courses, tutors, groups = _read_modules(xlsx_modules)
    rooms = _read_punggol_rooms(xlsx_resources)
    rooms.append(Room(
        id=VIRTUAL_ROOM_ID, name="Virtual (Online)",
        capacity=10_000, room_type=RoomType.VIRTUAL, zone="ONLINE",
    ))
    week_dates = {
        w: (SEMESTER_START_DATE + timedelta(days=(w - 1) * 7)).isoformat()
        for w in range(1, 14)
    }
    calendar = Calendar(teaching_weeks=list(range(1, 14)), week_dates=week_dates)
    return Universe(
        courses=courses, rooms=rooms, tutors=tutors, groups=groups,
        time_slots=_build_time_slots(), calendar=calendar,
    )


# ---------------------------------------------------------------------------
# Module sheet → Course / Activity / Tutor / Group
# ---------------------------------------------------------------------------

def _read_modules(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = list(wb["Module"].iter_rows(values_only=True))

    cur_prog = cur_size = cur_code = None
    raw = []
    placeholders: dict[tuple, dict] = {}   # (prog, code) -> last seen meta
    for r in rows[2:]:
        if not r or all(v is None for v in r):
            continue
        prog, size, code, activity, mode, weeks, s1, s1id, *_rest = r[:11]
        s2 = r[8] if len(r) > 8 else None
        s2id = r[9] if len(r) > 9 else None
        remarks = r[10] if len(r) > 10 else None
        if prog is not None:
            cur_prog = prog
        if size is not None:
            cur_size = size
        if code is not None:
            cur_code = code
        if activity is None:
            # Module code listed but no activity data — keep as a placeholder so
            # we can synthesise default sessions for it later.
            if cur_code:
                placeholders[(cur_prog, cur_code)] = dict(
                    prog=cur_prog, size=cur_size, code=cur_code,
                )
            continue
        raw.append(dict(prog=cur_prog, size=cur_size, code=cur_code,
                        activity=activity, mode=mode, weeks=weeks,
                        s1=s1, s1id=s1id, s2=s2, s2id=s2id, remarks=remarks))

    # Cross-cluster modules listed without Delivery Mode / Teaching Weeks
    # are silently skipped — per supervisor's instruction, we only schedule
    # modules with full activity data.  We still print which ones we skipped
    # so the planner has a clear audit trail.
    have_codes = {(r["prog"], r["code"]) for r in raw}
    skipped = []
    for (prog, code), meta in placeholders.items():
        if not prog or "DSC" not in str(prog).upper():
            continue
        if (prog, code) in have_codes:
            continue
        skipped.append(code)
    if skipped:
        import sys
        print(f"[loader] skipped {len(skipped)} module(s) without activity data: "
              f"{', '.join(sorted(set(skipped)))}", file=sys.stderr)

    raw = [r for r in raw if r["prog"] and "DSC" in str(r["prog"]).upper()]

    tutors_by_id: dict[str, Tutor] = {}

    aggregated: dict[tuple, dict] = {}
    for r in raw:
        try:
            atype = ActivityType.parse(str(r["activity"]))
            mode = DeliveryMode.parse(str(r["mode"]))
        except ValueError:
            continue
        key = (r["code"], atype, mode)
        slot = aggregated.setdefault(key, dict(
            code=r["code"], size=r["size"] or 0, prog=r["prog"],
            atype=atype, mode=mode, weeks=set(),
            s1=r["s1"], s1id=r["s1id"], remarks=r["remarks"],
        ))
        slot["weeks"].update(_parse_weeks(r["weeks"]))
        if r["s1id"] and not slot["s1id"]:
            slot["s1"], slot["s1id"] = r["s1"], r["s1id"]
        if r["size"]:
            slot["size"] = r["size"]

    by_code: dict[str, Course] = {}
    groups: list[Group] = []
    seen_groups: set[str] = set()

    for slot in aggregated.values():
        code = slot["code"]
        prog_label = str(slot["prog"])
        m = re.search(r"YR\s*(\d)", prog_label.upper())
        year = int(m.group(1)) if m else 0
        course = by_code.setdefault(code, Course(code=code, programme="DSC", year=year))

        atype = slot["atype"]
        mode = slot["mode"]
        size = slot["size"]
        tutor_name = slot["s1"] or "Unknown"
        tutor_id = slot["s1id"] or ("X_" + re.sub(r"[^A-Z0-9]", "", str(tutor_name).upper())[:24])
        tutors_by_id.setdefault(tutor_id, Tutor(id=tutor_id, name=tutor_name))

        weeks = sorted(slot["weeks"])
        duration = _default_duration(atype)

        if atype == ActivityType.LECTURE:
            gid = f"{code}/All"
            if gid not in seen_groups:
                groups.append(Group(id=gid, course_code=code, label="All", size=size))
                seen_groups.add(gid)
            fday, fstart = None, None
            course.activities.append(Activity(
                course_code=code, activity_type=atype, delivery_mode=mode,
                duration_slots=duration, weeks=weeks,
                tutor_id=tutor_id, group_id=gid, size=size,
                fixed_day=fday, fixed_start_index=fstart,
                notes=str(slot["remarks"] or ""),
            ))
        else:
            cap = _group_cap(atype)
            prefix = _group_prefix(atype)
            fday, fstart = None, None
            for label, gsize in _split_into_groups(size, cap, prefix):
                gid = f"{code}/{label}"
                if gid not in seen_groups:
                    groups.append(Group(id=gid, course_code=code, label=label, size=gsize))
                    seen_groups.add(gid)
                # Only pin the FIRST group to the named time — splitting tutorials
                # to different slots is the whole point of having multiple groups.
                pin_day = fday if label.endswith("1") else None
                pin_start = fstart if label.endswith("1") else None
                course.activities.append(Activity(
                    course_code=code, activity_type=atype, delivery_mode=mode,
                    duration_slots=duration, weeks=weeks,
                    tutor_id=tutor_id, group_id=gid, size=gsize,
                    fixed_day=pin_day, fixed_start_index=pin_start,
                    notes=str(slot["remarks"] or ""),
                ))

    courses = sorted(by_code.values(), key=lambda c: (c.year, c.code))
    tutors = sorted(tutors_by_id.values(), key=lambda t: t.name)
    return courses, tutors, groups


# ---------------------------------------------------------------------------
# Location sheet → Room
# ---------------------------------------------------------------------------

def _read_punggol_rooms(path: str) -> List[Room]:
    """Read the SIT Punggol Campus Court venue file.

    The file has columns:
        Location Name | Location Description | Capacity | Resource Type | Recording?

    Resource Type values map to our RoomType enum:
        Lectorial      → LECTURE_THEATRE
        Seminar Room   → SEMINAR_ROOM
        Laboratory     → LABORATORY
        Computer Room  → COMPUTER_LAB
        Auditorium     → LECTURE_THEATRE
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active   # single-sheet file
    rows = list(ws.iter_rows(values_only=True))[1:]   # skip header
    out: list[Room] = []
    seen: set[str] = set()
    type_map = {
        "lectorial": RoomType.LECTURE_THEATRE,
        "auditorium": RoomType.LECTURE_THEATRE,
        "seminar room": RoomType.SEMINAR_ROOM,
        "laboratory": RoomType.LABORATORY,
        "computer room": RoomType.COMPUTER_LAB,
    }
    for r in rows:
        if not r or not r[0]:
            continue
        name = str(r[0])
        cap = r[2]
        rtype = str(r[3] or "").strip().lower()
        try:
            cap = int(cap)
        except (TypeError, ValueError):
            continue
        if cap < 18:
            continue
        if name in seen:
            continue
        seen.add(name)
        # Building prefix (E2/E3/E5/E6) used as the zone label
        zone = name.split("-")[0] if "-" in name else "Punggol"
        out.append(Room(
            id=name, name=name, capacity=cap,
            room_type=type_map.get(rtype, RoomType.OTHER),
            zone=zone,
        ))
    out.sort(key=lambda r: (r.room_type.value, -r.capacity))
    return out


# ===========================================================================
# Single-file loader — reads from inputs.xlsx (the user-friendly customisable
# workbook).  Replaces the two-file SIT-supplied flow with one tab-per-entity
# workbook.
# ===========================================================================

_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}


def _date_to_week(token: str, semester_start) -> "int | None":
    """Convert a 'DD Mon [YYYY]' date inside `token` to a teaching-week number,
    counting from `semester_start` (Monday of Week 1).  Returns None if no date."""
    if not semester_start:
        return None
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3,9})\.?(?:\s+(\d{4}))?", token)
    if not m:
        return None
    day = int(m.group(1))
    mon = m.group(2)[:3].lower()
    if mon not in _MONTHS:
        return None
    year = int(m.group(3)) if m.group(3) else semester_start.year
    try:
        d = date(year, _MONTHS[mon], day)
    except ValueError:
        return None
    wk = (d - semester_start).days // 7 + 1
    return wk if wk >= 1 else None


def _parse_weeks_range(raw, semester_start=None) -> List[int]:
    """Parse a Teaching Weeks cell into a list of week numbers, tolerating the
    real-world SIT formats:
        '1-13'                       -> 1..13
        '1,2,3,8,9'                  -> those weeks
        '7 (14 Nov), 11 (11 Nov)'    -> 7, 11   (week number wins; note ignored)
        '10 Sep 2025, 8.30am-12pm'   -> 1       (date -> week; time token ignored)
        '22 Oct (Wed), 1.30pm-5pm'   -> 7       (date -> week; time token ignored)
    """
    if raw is None or str(raw).strip() == "":
        return list(range(1, 14))
    out: list[int] = []
    for tok in str(raw).split(','):
        tok = tok.strip()
        if not tok:
            continue
        # 1) Date with a month name (e.g. "10 Sep 2025", "22 Oct") -> week number.
        if re.match(r"^\d{1,2}\s+[A-Za-z]{3,}", tok):
            wk = _date_to_week(tok, semester_start)
            if wk is not None:
                out.append(wk)
            continue
        # 2) Pure time token (e.g. "8.30am-12pm", "1.30pm-5pm") -> ignore.
        if re.search(r"[ap]m", tok.lower()):
            continue
        # 3) Integer range "a-b".
        mr = re.match(r"^(\d+)\s*-\s*(\d+)$", tok)
        if mr:
            out.extend(range(int(mr.group(1)), int(mr.group(2)) + 1))
            continue
        # 4) Leading integer = week number (covers "7" and "7 (14 Nov)").
        ml = re.match(r"^(\d{1,2})\b", tok)
        if ml:
            out.append(int(ml.group(1)))
    return sorted(set(w for w in out if w >= 1)) or list(range(1, 14))


def _hhmm_to_slot(hhmm: str) -> int | None:
    """'10:00' or '14:30' → slot index (offset from DAY_START_HOUR)."""
    if not hhmm: return None
    s = str(hhmm).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m: return None
    h, mi = int(m.group(1)), int(m.group(2))
    if h < DAY_START_HOUR: return None
    return (h - DAY_START_HOUR) * 2 + (1 if mi >= 30 else 0)


_DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri"]


_MONTH_RE = re.compile(
    r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b', re.IGNORECASE)


def _extract_pin_from_weeks_cell(raw, semester_start=None) -> tuple:
    """Extract (day, start_slot, duration_slots) from a Teaching Weeks cell
    that contains a specific calendar date and time range, e.g.:
        '22 Oct (Wed), 1.30pm-5pm'  ->  ('Wed', 9, 7)   [13:30 with DAY_START=9, 3h30m]
        '10 Nov (Mon), 9am-1pm'     ->  ('Mon', 0, 8)
    Returns (None, None, None) when no calendar date is present.

    Strategy: LLM is tried first (handles any free-text format); regex is
    the fallback so the function still works without an API key.
    The day-of-week is always derived from the actual calendar date via
    date.weekday() to avoid annotation errors; the LLM day annotation is
    used only when date parsing fails.
    """
    if not raw:
        return None, None, None
    text = str(raw)

    # Skip plain week-number cells ("1-13", "1,3,5") — no calendar date present.
    has_text_date = bool(_MONTH_RE.search(text))
    has_iso_date = bool(re.search(r"\b\d{4}-\d{1,2}-\d{1,2}\b", text))
    if not has_text_date and not has_iso_date and not isinstance(raw, (date, datetime)):
        return None, None, None

    # ── Step 1: derive day from the calendar date (authoritative) ───────────
    day = None
    if isinstance(raw, datetime):
        d = raw.date()
        weekday = d.weekday()
        if weekday < 5:
            day = _DOW_NAMES[weekday]
    elif isinstance(raw, date):
        weekday = raw.weekday()
        if weekday < 5:
            day = _DOW_NAMES[weekday]
    else:
        m_iso = re.search(r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b", text)
        if m_iso:
            try:
                d = date(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
                weekday = d.weekday()
                if weekday < 5:
                    day = _DOW_NAMES[weekday]
            except ValueError:
                pass
        else:
            m_date = re.search(r"(\d{1,2})\s+([A-Za-z]{3,9})\.?(?:\s+(\d{4}))?", text)
            if m_date:
                day_num = int(m_date.group(1))
                mon_str = m_date.group(2)[:3].lower()
                if mon_str in _MONTHS:
                    ref_year = (int(m_date.group(3)) if m_date.group(3)
                                else (semester_start.year if semester_start
                                      else date.today().year))
                    try:
                        d = date(ref_year, _MONTHS[mon_str], day_num)
                        weekday = d.weekday()
                        if weekday < 5:
                            day = _DOW_NAMES[weekday]
                    except ValueError:
                        pass

    # ── Step 2: extract time range ──────────────────────────────────────────
    start_slot = None
    dur_slots  = None

    # Primary: LLM (handles any free-text time format)
    try:
        from .remarks_parser import parse_weeks_cell_llm
        llm = parse_weeks_cell_llm(text)
        if llm:
            sh = llm.get("start_hour")
            sm = int(llm.get("start_min") or 0)
            eh = llm.get("end_hour")
            em = int(llm.get("end_min") or 0)
            # Use LLM day only when calendar-date derivation failed
            if day is None and llm.get("pin_day") in ("Mon","Tue","Wed","Thu","Fri"):
                day = llm["pin_day"]
            if sh is not None and sh >= DAY_START_HOUR:
                start_slot = (sh - DAY_START_HOUR) * 2 + (1 if sm >= 30 else 0)
                if eh is not None and (eh * 60 + em) > (sh * 60 + sm):
                    dur_slots = ((eh * 60 + em) - (sh * 60 + sm)) // 30
    except Exception:
        pass

    # Fallback: regex time range (works without API key)
    if start_slot is None:
        m_range = re.search(
            r"(\d{1,2})(?:[.:](\d{2}))?\s*(am|pm)\s*-\s*(\d{1,2})(?:[.:](\d{2}))?\s*(am|pm)",
            text, re.IGNORECASE)
        if m_range:
            sh = int(m_range.group(1)); sm = int(m_range.group(2) or 0)
            ssuf = m_range.group(3).lower()
            eh = int(m_range.group(4)); em = int(m_range.group(5) or 0)
            esuf = m_range.group(6).lower()
            if ssuf == "pm" and sh < 12: sh += 12
            if ssuf == "am" and sh == 12: sh = 0
            if esuf == "pm" and eh < 12: eh += 12
            if esuf == "am" and eh == 12: eh = 0
            start_total = sh * 60 + sm
            end_total   = eh * 60 + em
            if sh >= DAY_START_HOUR and end_total > start_total:
                start_slot = (sh - DAY_START_HOUR) * 2 + (1 if sm >= 30 else 0)
                dur_slots  = (end_total - start_total) // 30
        else:
            # Start time only (no end time → no duration)
            m_time = re.search(r"(\d{1,2})[.:](\d{2})\s*(am|pm)", text, re.IGNORECASE)
            if not m_time:
                m_time = re.search(r"\b(\d{1,2})\s*(am|pm)\b", text, re.IGNORECASE)
            if m_time:
                groups = m_time.groups()
                hour   = int(groups[0])
                minute = int(groups[1]) if len(groups) > 2 else 0
                suffix = groups[-1].lower()
                if suffix == "pm" and hour < 12: hour += 12
                if suffix == "am" and hour == 12: hour = 0
                if hour >= DAY_START_HOUR:
                    start_slot = (hour - DAY_START_HOUR) * 2 + (1 if minute >= 30 else 0)

    return day, start_slot, dur_slots




def _apply_settings(wb) -> None:
    """Read Settings tab and override module-level constants accordingly.

    The Settings sheet uses a key-value layout: column A is the label,
    column B is the value.  We look up labels case-insensitively and
    leave defaults in place when a row is missing or blank.
    """
    if "Settings" not in wb.sheetnames:
        return
    ws = wb["Settings"]
    kv: dict[str, object] = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and row[1] is not None:
            kv[str(row[0]).strip().lower()] = row[1]

    global DAY_START_HOUR, DAY_END_HOUR, SLOT_MIN, DAYS
    global PERIOD_STARTS_2H, PERIOD_STARTS_3H
    global DEFAULT_TUTORIAL_CAP, DEFAULT_LAB_CAP, DEFAULT_WORKSHOP_CAP

    def _hh_to_slot(hh: str, day_start: int) -> int:
        m = re.match(r"\s*(\d{1,2})(?::(\d{2}))?", str(hh))
        if not m: return -1
        h = int(m.group(1)); mi = int(m.group(2) or 0)
        return (h - day_start) * 2 + (1 if mi >= 30 else 0)

    try:
        if "day start hour" in kv:        DAY_START_HOUR = int(kv["day start hour"])
        if "day end hour" in kv:          DAY_END_HOUR   = int(kv["day end hour"])
        if "slot duration (minutes)" in kv: SLOT_MIN     = int(kv["slot duration (minutes)"])
        if "working days" in kv:
            DAYS = [d.strip() for d in str(kv["working days"]).split(",") if d.strip()]
    except (TypeError, ValueError):
        pass

    # Canonical periods — parse "09:00, 12:00, 14:00, 16:00" relative to new DAY_START_HOUR
    if "2-hour blocks (lec/tut)" in kv:
        PERIOD_STARTS_2H = [_hh_to_slot(t, DAY_START_HOUR)
                            for t in str(kv["2-hour blocks (lec/tut)"]).split(",")]
        PERIOD_STARTS_2H = [s for s in PERIOD_STARTS_2H if s >= 0]
    if "3-hour blocks (lab/ws)" in kv:
        PERIOD_STARTS_3H = [_hh_to_slot(t, DAY_START_HOUR)
                            for t in str(kv["3-hour blocks (lab/ws)"]).split(",")]
        PERIOD_STARTS_3H = [s for s in PERIOD_STARTS_3H if s >= 0]

    # Group caps
    try:
        if "tutorial cap" in kv: DEFAULT_TUTORIAL_CAP = int(kv["tutorial cap"])
        if "lab cap"      in kv: DEFAULT_LAB_CAP      = int(kv["lab cap"])
        if "workshop cap" in kv: DEFAULT_WORKSHOP_CAP = int(kv["workshop cap"])
    except (TypeError, ValueError):
        pass

    # Soft constraint weights — these live in constraint_engine.WEIGHTS
    try:
        from . import constraint_engine as _ce
        if "s1 mode switch (online\u2194f2f)" in kv: _ce.WEIGHTS["S1_mode_switch"] = int(kv["s1 mode switch (online\u2194f2f)"])
        for k_in, k_out in [
            ("s2 tutor idle gap >2h", "S2_tutor_gap"),
            ("s3 group long block >4h", "S3_long_block"),
            ("s4 short campus day (1-2h)", "S4_short_day"),
            ("s5 online not mon/tue", "S5_online_day"),
        ]:
            if k_in in kv:
                _ce.WEIGHTS[k_out] = int(kv[k_in])
        # S1 may have unicode arrow that varies
        for k in kv:
            if k.startswith("s1 mode switch"):
                _ce.WEIGHTS["S1_mode_switch"] = int(kv[k])
                break
    except (ImportError, TypeError, ValueError):
        pass


def load_from_inputs(path: str) -> Universe:
    """Load Universe from a single user-friendly inputs.xlsx workbook.

    Tabs expected: Modules, Tutors, Rooms, Calendar, Settings.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Apply Settings overrides BEFORE deriving time slots ------------
    _apply_settings(wb)

    settings = wb["Settings"]
    s = {row[0]: row[1] for row in settings.iter_rows(values_only=True) if row and row[0]}

    # ---- Tutors ---------------------------------------------------------
    tutors_ws = wb["Tutors"]
    tutors_by_id: dict[str, Tutor] = {}
    for r in list(tutors_ws.iter_rows(values_only=True))[1:]:
        if not r or not r[0]: continue
        tid, tname = str(r[0]), str(r[1] or r[0])
        tutors_by_id[tid] = Tutor(id=tid, name=tname)

    # ---- Rooms ----------------------------------------------------------
    rooms_ws = wb["Rooms"]
    rooms: list[Room] = []
    for r in list(rooms_ws.iter_rows(values_only=True))[1:]:
        if not r or not r[0]: continue
        rid = str(r[0])
        try: cap = int(r[1])
        except: continue
        rtype_str = str(r[2] or "other").strip().lower()
        type_map = {
            "lecture_theatre": RoomType.LECTURE_THEATRE,
            "lectorial":       RoomType.LECTURE_THEATRE,
            "auditorium":      RoomType.LECTURE_THEATRE,
            "seminar_room":    RoomType.SEMINAR_ROOM,
            "seminar room":    RoomType.SEMINAR_ROOM,
            "laboratory":      RoomType.LABORATORY,
            "computer_lab":    RoomType.COMPUTER_LAB,
            "computer room":   RoomType.COMPUTER_LAB,
            "virtual":         RoomType.VIRTUAL,
            "other":           RoomType.OTHER,
        }
        rooms.append(Room(
            id=rid, name=rid, capacity=cap,
            room_type=type_map.get(rtype_str, RoomType.OTHER),
            zone=str(r[3] or ""),
        ))

    # ---- Modules → Courses + Activities + Groups ------------------------
    modules_ws = wb["Modules"]
    by_code: dict[str, Course] = {}
    groups: list[Group] = []
    seen_groups: set[str] = set()
    for r in list(modules_ws.iter_rows(values_only=True))[1:]:
        if not r or not r[0]: continue
        code = str(r[0])
        prog = str(r[1] or "DSC")
        try: year = int(r[2] or 0)
        except: year = 0
        try: size = int(r[3] or 0)
        except: continue
        atype_raw = str(r[4] or "").strip()
        mode_raw = str(r[5] or "").strip()
        weeks_raw = r[6]
        try: dur_h = float(r[7] or 2)
        except: dur_h = 2
        tutor_id = str(r[8] or "UNKNOWN")
        pin_day = str(r[9] or "").strip() or None
        pin_start = _hhmm_to_slot(str(r[10] or ""))
        notes = str(r[11] or "")

        try:
            atype = ActivityType.parse(atype_raw)
            mode = DeliveryMode.parse(mode_raw)
        except ValueError:
            continue

        course = by_code.setdefault(code, Course(code=code, programme=prog, year=year))
        weeks = _parse_weeks_range(weeks_raw)
        duration_slots = int(round(dur_h * 60 / SLOT_MIN))

        # Lectures and "Other" (cohort-wide events such as DSC3002A's
        # industry engagement, or any non-tutorial gathering) share a
        # single "All" group — they're not split into sub-groups.
        if atype in (ActivityType.LECTURE, ActivityType.OTHER):
            gid = f"{code}/All"
            if gid not in seen_groups:
                groups.append(Group(id=gid, course_code=code, label="All", size=size))
                seen_groups.add(gid)
            course.activities.append(Activity(
                course_code=code, activity_type=atype, delivery_mode=mode,
                duration_slots=duration_slots, weeks=weeks,
                tutor_id=tutor_id, group_id=gid, size=size,
                fixed_day=pin_day, fixed_start_index=pin_start,
                notes=notes,
            ))
        else:
            cap = _group_cap(atype)
            prefix = _group_prefix(atype)
            for label, gsize in _split_into_groups(size, cap, prefix):
                gid = f"{code}/{label}"
                if gid not in seen_groups:
                    groups.append(Group(id=gid, course_code=code, label=label, size=gsize))
                    seen_groups.add(gid)
                # Pin only the FIRST group; others stay free
                pin_d = pin_day if label.endswith("1") else None
                pin_s = pin_start if label.endswith("1") else None
                course.activities.append(Activity(
                    course_code=code, activity_type=atype, delivery_mode=mode,
                    duration_slots=duration_slots, weeks=weeks,
                    tutor_id=tutor_id, group_id=gid, size=gsize,
                    fixed_day=pin_d, fixed_start_index=pin_s,
                    notes=notes,
                ))

    # ---- Calendar -------------------------------------------------------
    cal_ws = wb["Calendar"]
    cal_kv = {row[0]: row[1] for row in cal_ws.iter_rows(values_only=True) if row and row[0]}
    sem_start_raw = cal_kv.get("Semester start date")
    if sem_start_raw:
        if hasattr(sem_start_raw, "year"):
            sem_start = sem_start_raw if isinstance(sem_start_raw, date) else date.fromisoformat(str(sem_start_raw))
        else:
            sem_start = date.fromisoformat(str(sem_start_raw)[:10])
    else:
        sem_start = SEMESTER_START_DATE
    weeks_list = _parse_weeks_range(cal_kv.get("Teaching weeks") or "1-13")
    week_dates = {w: (sem_start + timedelta(days=(w - 1) * 7)).isoformat() for w in weeks_list}
    calendar = Calendar(teaching_weeks=weeks_list, week_dates=week_dates)

    courses = sorted(by_code.values(), key=lambda c: (c.year, c.code))
    return Universe(
        courses=courses,
        rooms=rooms,
        tutors=sorted(tutors_by_id.values(), key=lambda t: t.name),
        groups=groups,
        time_slots=_build_time_slots(),
        calendar=calendar,
    )


# ===========================================================================
# Worksheet loader — reads the SIT-style "Module" layout *as supplied*:
#
#   Prog/Yr | Class Size | Module Code | Activity | Delivery Mode |
#   Teaching Weeks | Staff 1 | Staff ID 1 | Staff 2 | Staff ID 2 | Remarks
#
# This is the layout the planner hands over.  It is programme-agnostic
# (ASE, DSC, INF, ...), the planner provides the staff columns, and the
# solver assigns time + room for every activity.  There are no Pin / Duration
# columns, so every class is placed automatically (durations come from the
# activity-type defaults; an optional time pin can still be written in the
# Remarks column, e.g. "Monday, 10am-12pm").
# ===========================================================================

def _read_inputs_rooms(path: str) -> List[Room]:
    """Pull the curated room list (incl. the VIRTUAL room) from an inputs.xlsx
    Rooms tab.  Used so a bare worksheet doesn't need its own venue file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Rooms" not in wb.sheetnames:
        raise ValueError(f"{path} has no 'Rooms' tab to source venues from")
    type_map = {
        "lecture_theatre": RoomType.LECTURE_THEATRE,
        "seminar_room":    RoomType.SEMINAR_ROOM,
        "laboratory":      RoomType.LABORATORY,
        "computer_lab":    RoomType.COMPUTER_LAB,
        "virtual":         RoomType.VIRTUAL,
        "other":           RoomType.OTHER,
    }
    rooms: list[Room] = []
    for r in list(wb["Rooms"].iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        rid = str(r[0])
        try:
            cap = int(r[1])
        except (TypeError, ValueError):
            continue
        rtype = str(r[2] or "other").strip().lower()
        rooms.append(Room(
            id=rid, name=rid, capacity=cap,
            room_type=type_map.get(rtype, RoomType.OTHER),
            zone=str(r[3] or ""),
        ))
    if not any(rm.is_virtual for rm in rooms):
        rooms.append(Room(id=VIRTUAL_ROOM_ID, name="Virtual (Online)",
                          capacity=10_000, room_type=RoomType.VIRTUAL, zone="ONLINE"))
    return rooms


def _prog_and_year(prog_label: str):
    """'ASE/Yr 1' -> ('ASE', 1);  'DSC/YR 2' -> ('DSC', 2)."""
    label = str(prog_label or "")
    prefix = re.split(r"[/\s]+", label.strip())[0].upper() if label.strip() else "GEN"
    m = re.search(r"(\d+)", label)
    year = int(m.group(1)) if m else 0
    return prefix, year


def _prog_year_from_code(code: str):
    """Derive (programme, year) from a module code, e.g. 'DSC1001' -> ('DSC', 1),
    'MET2602' -> ('MET', 2).  Lets the planner omit the Prog/Yr column entirely."""
    m = re.match(r"\s*([A-Za-z]+)\s*(\d)", str(code or ""))
    if not m:
        return ("GEN", 0)
    return (m.group(1).upper(), int(m.group(2)))


def _split_weeks(weeks, n):
    """Partition a sorted week list into n contiguous, near-even blocks.
    Used when the solver auto-assigns several eligible tutors to one activity
    (tutor A takes the first block of weeks, B the next, and so on)."""
    weeks = sorted(weeks)
    if n <= 1 or len(weeks) <= 1:
        return [weeks]
    n = min(n, len(weeks))
    k, m = divmod(len(weeks), n)
    blocks, i = [], 0
    for x in range(n):
        size = k + (1 if x < m else 0)
        blocks.append(weeks[i:i + size])
        i += size
    return blocks


def _read_eligibility(wb) -> dict:
    """Read an 'Eligibility' tab: Module Code | Activity | Eligible Staff IDs.

    Returns {(code, activity_lower or None): [staff_id, ...]}.  A blank Activity
    means the list applies to every activity of that module.  For module rows
    that leave the staff columns blank, the solver auto-assigns staff drawn ONLY
    from these eligibility lists (and may split the weeks across them)."""
    target = None
    for s in wb.sheetnames:
        if s.strip().lower() in ("eligibility", "staff eligibility", "eligible staff"):
            target = s
            break
    elig: dict = {}
    if not target:
        return elig
    for r in list(wb[target].iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        code = str(r[0]).strip()
        act = (str(r[1]).strip().lower() if len(r) > 1 and r[1] else None)
        ids_cell = r[2] if len(r) > 2 else None
        ids = [x.strip() for x in re.split(r"[,;]", str(ids_cell or "")) if x.strip()]
        if ids:
            elig[(code, act)] = ids
    return elig


def _read_activity_modes(wb) -> dict:
    """Read an 'ActivityModes' tab mapping each Activity to its permitted
    delivery modes (Activity | Mode 1 | Mode 2 | Mode 3 ...).

    Returns {ActivityType: set[DeliveryMode]}.  Activity types NOT listed are
    treated as unrestricted (any mode allowed).  This is the SIT rule that,
    e.g., a Lecture may only be Online (sync/async), never f2f.
    """
    target = None
    for s in wb.sheetnames:
        if s.strip().lower().replace("_", " ") in ("activitymodes", "activity modes", "activity types"):
            target = s
            break
    rules: dict = {}
    if not target:
        return rules
    rows = list(wb[target].iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        atype = ActivityType.parse(str(r[0]))
        modes = set()
        for cell in r[1:]:
            if cell is None or str(cell).strip() == "":
                continue
            try:
                modes.add(DeliveryMode.parse(str(cell)))
            except ValueError:
                pass
        if modes:
            rules[atype] = modes
    return rules


def load_from_worksheet(worksheet_path: str,
                        rooms_inputs: str | None = None,
                        sheet_name: str | None = None,
                        semester_start: str = "2025-09-08",
                        teaching_weeks: str = "1-13",
                        ignore_remarks: bool = False) -> Universe:
    """Build a Universe straight from the supplied SIT 'Module' worksheet.

    Parameters
    ----------
    worksheet_path : the workbook holding the module/activity/staff layout.
    rooms_inputs   : path to an inputs.xlsx whose Rooms tab supplies venues.
                     Defaults to ../inputs.xlsx next to the worksheet.
    sheet_name     : sheet to read; auto-detected (the one with a
                     'Module Code' header) when omitted.
    """
    wb = openpyxl.load_workbook(worksheet_path, data_only=True)

    # When this workbook also carries Settings / Tutors / Rooms / Calendar
    # tabs (i.e. it is the centralised single-file input) use them, so the
    # whole timetable is driven from one file.
    _apply_settings(wb)
    tutor_names_override: dict[str, str] = {}
    if "Tutors" in wb.sheetnames:
        for tr in list(wb["Tutors"].iter_rows(values_only=True))[1:]:
            if tr and tr[0]:
                tutor_names_override[str(tr[0]).strip()] = str(tr[1] or tr[0]).strip()
    allowed_modes = _read_activity_modes(wb)
    mode_warnings: list[str] = []
    eligibility = _read_eligibility(wb)
    staff_warnings: list[str] = []
    _sem_start = date.fromisoformat(str(semester_start)[:10])
    if "Calendar" in wb.sheetnames:
        _ck = {row[0]: row[1] for row in wb["Calendar"].iter_rows(values_only=True) if row and row[0]}
        _rs = _ck.get("Semester start date")
        if _rs is not None and hasattr(_rs, "year"):
            _sem_start = _rs if isinstance(_rs, date) else date.fromisoformat(str(_rs)[:10])
        elif _rs:
            _sem_start = date.fromisoformat(str(_rs)[:10])

    # ---- locate the right sheet + header row ----------------------------
    def _has_module_header(ws):
        for r in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            if r and any(str(c).strip().lower() == "module code" for c in r if c):
                return True
        return False

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = next((wb[s] for s in wb.sheetnames if _has_module_header(wb[s])), wb.active)

    rows = list(ws.iter_rows(values_only=True))
    header_idx = 0
    for i, r in enumerate(rows):
        if r and any(str(c).strip().lower() == "module code" for c in r if c):
            header_idx = i
            break

    # ---- read rows, carrying down merged Prog / Size / Code -------------
    # Build a header-name -> column-index map so column ORDER in Module doesn't matter
    _cmap = {}
    for _ci, _hv in enumerate(rows[header_idx]):
        if _hv is not None:
            _cmap[str(_hv).strip().lower()] = _ci

    # Detect ALL "Staff N" / "Staff ID N" pairs dynamically so the worksheet
    # can carry Staff 1..Staff K without having to teach the loader each one.
    _staff_col_pairs = []
    for _hdr_name, _ci in _cmap.items():
        _m = re.match(r"^staff\s+(\d+)$", _hdr_name)
        if _m:
            _n = int(_m.group(1))
            _staff_col_pairs.append((_n, _ci, _cmap.get(f"staff id {_n}")))
    _staff_col_pairs.sort(key=lambda x: x[0])
    if not _staff_col_pairs:
        # Fall back to the legacy fixed pair if the worksheet doesn't have
        # explicitly numbered staff columns.
        _staff_col_pairs = [(1, _cmap.get("staff 1"), _cmap.get("staff id 1")),
                            (2, _cmap.get("staff 2"), _cmap.get("staff id 2"))]

    cur_prog = cur_size = cur_code = None
    raw = []
    for r in rows[header_idx + 1:]:
        if not r or all(v is None for v in r):
            continue
        cells = list(r)
        def _g(name, default=None):
            i = _cmap.get(name)
            return cells[i] if (i is not None and i < len(cells)) else default
        prog     = _g("prog/yr")
        size     = _g("class size")
        code     = _g("module code")
        activity = _g("activity")
        mode     = _g("delivery mode")
        weeks    = _g("teaching weeks")
        remarks  = _g("remarks")
        subgroups= _g("subgroups")
        # All Staff N / Staff ID N pairs that have any content on THIS row,
        # preserving their column order (Staff 1, Staff 2, Staff 3, ...).
        # TBD placeholders left behind by a previous write-back (Staff column
        # value "TBD (unassigned)" / Staff ID starting with "TBD_") are
        # ignored — otherwise the loader would treat the placeholder as a
        # real teacher and consume one of the subgroup slots.
        row_staff = []
        for _n, _name_ci, _id_ci in _staff_col_pairs:
            _nm = cells[_name_ci] if (_name_ci is not None and _name_ci < len(cells)) else None
            _idv = cells[_id_ci] if (_id_ci is not None and _id_ci < len(cells)) else None
            _ns = (str(_nm).strip() if _nm not in (None, "") else "")
            _is = (str(_idv).strip() if _idv not in (None, "") else "")
            if _is.upper().startswith("TBD") or _ns.upper().startswith("TBD"):
                continue
            if _ns or _is:
                row_staff.append((_ns, _is))
        if prog is not None:
            cur_prog = prog
        if size is not None:
            cur_size = size
        if code is not None:
            cur_code = code
        if activity is None:        # module listed without activity data — skip
            continue
        evening = bool(_g("start at 7pm?"))
        raw.append(dict(prog=cur_prog, size=cur_size, code=cur_code,
                        activity=activity, mode=mode, weeks=weeks,
                        row_staff=row_staff,
                        remarks=remarks, subgroups=subgroups,
                        evening=evening))

    # ---- aggregate identical (code, activity, mode) across week rows ----
    #
    # Key change vs older versions: we no longer fold the Staff ID into the
    # aggregation key.  Two rows for the SAME (module, activity, delivery
    # mode) with DIFFERENT staff now end up in ONE slot whose `rows` list
    # records each row's own staff + week range.  That single slot is what
    # lets us distinguish:
    #   * parallel-subgroup teaching (subgroups>=2 AND #distinct staff
    #     matches) — e.g. ENG1002 Lecture with Kyrin Liong on weeks 1–6 and
    #     Xiang Ning on weeks 8–13 means "each prof teaches ONE half of the
    #     cohort for ALL 12 teaching weeks" (the user's rule), and
    #   * legacy week-split teaching (subgroups blank/1) — e.g. DSC1001
    #     Lecture taught by DAVID weeks 1–7 and YANG weeks 8–13 means "the
    #     SAME single group sees DAVID for the first half and YANG for the
    #     second half".
    aggregated: dict[tuple, dict] = {}
    for r in raw:
        try:
            atype = ActivityType.parse(str(r["activity"]))
            mode = DeliveryMode.parse(str(r["mode"]))
        except ValueError:
            continue
        key = (r["code"], atype, mode)
        slot = aggregated.setdefault(key, dict(
            code=r["code"], size=r["size"] or 0, prog=r["prog"],
            atype=atype, mode=mode, weeks=set(),
            remarks=r["remarks"], subgroups=r.get("subgroups"),
            rows=[], is_evening=False,
            raw_weeks_cell=r["weeks"],   # preserved for day+time pin extraction
            _has_empty_weeks=False,      # set True when any row's Teaching Weeks cell is blank
            _has_explicit_weeks=False,   # set True when any row has explicit week values
        ))
        row_weeks = set(_parse_weeks_range(r["weeks"], _sem_start))
        if r["weeks"] is None or str(r["weeks"]).strip() == "":
            slot["_has_empty_weeks"] = True
        else:
            slot["_has_explicit_weeks"] = True
        slot["weeks"].update(row_weeks)
        slot["rows"].append({
            "staff": list(r.get("row_staff") or []),
            "weeks": row_weeks,
        })
        if r.get("evening"):
            slot["is_evening"] = True
        if r["size"]:
            slot["size"] = r["size"]
        if r.get("subgroups") not in (None, "") and slot.get("subgroups") in (None, ""):
            slot["subgroups"] = r.get("subgroups")
        if r["remarks"] and not slot["remarks"]:
            slot["remarks"] = r["remarks"]

    by_code: dict[str, Course] = {}
    tutors_by_id: dict[str, Tutor] = {}
    groups: list[Group] = []
    seen_groups: set[str] = set()
    roster_ids = list(tutor_names_override.keys())
    assign_load: dict = {}

    def _staff_pair_to_tid(_ns: str, _is: str):
        """Map a (name, id) pair from a Staff column to (tutor_id, tutor_name),
        honouring tutor-name overrides from the Tutors tab."""
        _tid = _is or ("X_" + re.sub(r"[^A-Z0-9]", "", (_ns or _is).upper())[:24])
        _tname = tutor_names_override.get(_tid, _ns or _is or _tid)
        return _tid, _tname

    for slot in aggregated.values():
        code = slot["code"]
        if slot["prog"]:
            prog_prefix, year = _prog_and_year(slot["prog"])
        else:
            prog_prefix, year = _prog_year_from_code(code)
        course = by_code.setdefault(code, Course(code=code, programme=prog_prefix, year=year))

        atype, mode, size = slot["atype"], slot["mode"], slot["size"]
        if not size:
            size = DEFAULT_CLASS_SIZE
            staff_warnings.append(f"{code} {atype.value}: no Class Size given -> assumed {DEFAULT_CLASS_SIZE}")
        if atype in allowed_modes and mode not in allowed_modes[atype]:
            _ok = ", ".join(sorted(mm.value for mm in allowed_modes[atype]))
            mode_warnings.append(
                f"{slot['code']} {atype.value}: delivery mode '{mode.value}' "
                f"is not permitted (allowed: {_ok})")

        all_weeks = sorted(slot["weeks"])
        duration = _default_duration(atype)
        is_evening = bool(slot.get("is_evening", False))
        if ignore_remarks:
            fday, fstart = None, None
        else:
            fday, fstart = None, None
            if slot["remarks"]:
                from .remarks_parser import parse_remarks_llm
                fday, fstart = parse_remarks_llm(slot["remarks"])
            # Last resort: extract day+time embedded in the Teaching Weeks cell.
            # "22 Oct (Wed), 1.30pm-5pm" pins to Wed 13:30 with duration 3.5h;
            # the day is computed from the date, not the annotation.
            pin_day, pin_start, fdur = _extract_pin_from_weeks_cell(
                slot.get("raw_weeks_cell"), _sem_start)
            if fday is None and pin_day is not None:
                fday = pin_day
            if fstart is None and pin_start is not None:
                fstart = pin_start
            if fdur is not None:
                duration = fdur   # override default with cell-derived duration

        # ---- distinct staff across every row of this (code, atype, mode) ----
        distinct_staff: list = []
        seen_sids: set = set()
        for _row in slot["rows"]:
            for _ns, _is in _row["staff"]:
                _tid, _tname = _staff_pair_to_tid(_ns, _is)
                if _tid in seen_sids:
                    continue
                seen_sids.add(_tid)
                distinct_staff.append((_tid, _tname))
        # Largest single-row staff count — used so that the legacy
        # "Staff 1 + Staff 2 on ONE row = 2 parallel subgroups" default
        # still applies when subgroups column is blank.
        _max_row_staff = max((len(r["staff"]) for r in slot["rows"]), default=0)

        # ---- decide the per-subgroup staff plan ----
        # Number of subgroups:
        #   1) admin-set "Subgroups" column wins,
        #   2) otherwise the legacy default (the widest staff-row's count, or 1).
        _n_admin = slot.get("subgroups")
        try: _n_admin = int(_n_admin) if _n_admin not in (None, "") else None
        except (TypeError, ValueError): _n_admin = None
        _n_sub = _n_admin if _n_admin else max(1, _max_row_staff or 1)

        # ---- top up from Eligibility when the planner under-named ----
        # If Subgroups > #named staff and the Eligibility tab lists more
        # eligible people, append them to distinct_staff (skipping anyone
        # already named).  This turns the Eligibility list into a real
        # "supplement" pool, not just a fallback for fully-blank rows.
        if _n_sub > len(distinct_staff):
            _elig_ids = (eligibility.get((code, atype.value.lower()))
                         or eligibility.get((code, None)) or [])
            for _eid in _elig_ids:
                if _eid in seen_sids:
                    continue
                seen_sids.add(_eid)
                distinct_staff.append((_eid, tutor_names_override.get(_eid, _eid)))
                if len(distinct_staff) >= _n_sub:
                    break

        # PARALLEL-SUBGROUP MODE: each subgroup is taught by ONE staff for
        # the WHOLE union of teaching weeks.  Triggered when subgroups>=2
        # AND we know exactly that many distinct staff (so the mapping
        # subgroup_k -> distinct_staff[k] is unambiguous).  This handles
        #   * the ENG1002 case (2 rows × 1 staff, Subgroups=2), and
        #   * the legacy "single row carries Staff 1, Staff 2, ..." case.
        parallel_mode = (_n_sub >= 2 and len(distinct_staff) == _n_sub)

        # WEEK-SPLIT MODE: keep the old per-row emission, where each row
        # produces its own activity with its own staff + own week range
        # (the SAME group sees different tutors in different weeks).  This
        # is the DSC1001 lecture pattern and the DSC1001 tutorial pattern.
        any_named_staff = any(r["staff"] for r in slot["rows"])

        if parallel_mode:
            # Emit N subgroup activities, each glued to one staff for the
            # UNION of teaching weeks.
            prefix = ("G" if atype in (ActivityType.LECTURE, ActivityType.OTHER)
                      else _group_prefix(atype))
            for k, (label, gsize) in enumerate(_split_into_n_groups(size, _n_sub, prefix)):
                gid = f"{code}/{label}"
                if gid not in seen_groups:
                    groups.append(Group(id=gid, course_code=code, label=label, size=gsize))
                    seen_groups.add(gid)
                t_id, t_nm = distinct_staff[k]
                tutors_by_id.setdefault(t_id, Tutor(id=t_id, name=t_nm))
                pin_day = fday if label.endswith("1") else None
                pin_start = fstart if label.endswith("1") else None
                course.activities.append(Activity(
                    course_code=code, activity_type=atype, delivery_mode=mode,
                    duration_slots=duration, weeks=all_weeks,
                    tutor_id=t_id, group_id=gid, size=gsize,
                    fixed_day=pin_day, fixed_start_index=pin_start,
                    is_evening=is_evening,
                    notes=str(slot["remarks"] or ""), co_tutor_ids=[],
                    weeks_from_default=slot.get("_has_empty_weeks", False) and not slot.get("_has_explicit_weeks", False)))
        elif any_named_staff:
            # Per-row emission preserves week-split (DSC1001 lecture pattern,
            # DSC1001 tutorial pattern).  Each row gets ONE activity per
            # subgroup using THAT row's staff/weeks.
            for _row in slot["rows"]:
                if not _row["staff"]:
                    continue
                row_pairs = [_staff_pair_to_tid(ns, is_) for ns, is_ in _row["staff"]]
                wk = sorted(_row["weeks"])

                def _tutor_for(k, _row_pairs=row_pairs):
                    if k < len(_row_pairs):
                        return _row_pairs[k]
                    return (f"TBD_{code}_{atype.value}_{k+1}", "TBD (unassigned)")

                if atype in (ActivityType.LECTURE, ActivityType.OTHER) and _n_sub == 1:
                    gid = f"{code}/All"
                    if gid not in seen_groups:
                        groups.append(Group(id=gid, course_code=code, label="All", size=size))
                        seen_groups.add(gid)
                    t_id, t_nm = _tutor_for(0)
                    tutors_by_id.setdefault(t_id, Tutor(id=t_id, name=t_nm))
                    _co_pairs = row_pairs[1:]
                    for _c_id, _c_nm in _co_pairs:
                        tutors_by_id.setdefault(_c_id, Tutor(id=_c_id, name=_c_nm))
                    course.activities.append(Activity(
                        course_code=code, activity_type=atype, delivery_mode=mode,
                        duration_slots=duration, weeks=wk,
                        tutor_id=t_id, group_id=gid, size=size,
                        fixed_day=fday, fixed_start_index=fstart,
                        is_evening=is_evening,
                        notes=str(slot["remarks"] or ""), co_tutor_ids=[p[0] for p in _co_pairs],
                        weeks_from_default=slot.get("_has_empty_weeks", False) and not slot.get("_has_explicit_weeks", False)))
                else:
                    prefix = ("G" if atype in (ActivityType.LECTURE, ActivityType.OTHER)
                              else _group_prefix(atype))
                    for k, (label, gsize) in enumerate(_split_into_n_groups(size, _n_sub, prefix)):
                        gid = f"{code}/{label}"
                        if gid not in seen_groups:
                            groups.append(Group(id=gid, course_code=code, label=label, size=gsize))
                            seen_groups.add(gid)
                        t_id, t_nm = _tutor_for(k)
                        tutors_by_id.setdefault(t_id, Tutor(id=t_id, name=t_nm))
                        pin_day = fday if label.endswith("1") else None
                        pin_start = fstart if label.endswith("1") else None
                        course.activities.append(Activity(
                            course_code=code, activity_type=atype, delivery_mode=mode,
                            duration_slots=duration, weeks=wk,
                            tutor_id=t_id, group_id=gid, size=gsize,
                            fixed_day=pin_day, fixed_start_index=pin_start,
                            is_evening=is_evening,
                            notes=str(slot["remarks"] or ""), co_tutor_ids=[],
                            weeks_from_default=slot.get("_has_empty_weeks", False) and not slot.get("_has_explicit_weeks", False)))
        else:
            # NO planner-named staff anywhere — fall back to Eligibility list,
            # then the roster, then TBD.  This branch handles the auto-assign
            # paths and uses the UNION of weeks across all rows.
            eligible = (eligibility.get((code, atype.value.lower()))
                        or eligibility.get((code, None)))
            staff_plan: list = []
            if eligible:
                for tid, blk in zip(eligible, _split_weeks(all_weeks, len(eligible))):
                    if blk:
                        staff_plan.append((tid, tutor_names_override.get(tid, tid), blk))
            elif roster_ids:
                tid = min(roster_ids, key=lambda t: assign_load.get(t, 0))
                assign_load[tid] = assign_load.get(tid, 0) + len(all_weeks)
                staff_plan.append((tid, tutor_names_override.get(tid, tid), all_weeks))
            else:
                tid = f"TBD_{code}_{atype.value}"
                staff_plan.append((tid, "TBD (unassigned)", all_weeks))
                staff_warnings.append(
                    f"{code} {atype.value}: no staff, no eligibility, empty roster -> TBD")

            for tid, tname, wk in staff_plan:
                tutors_by_id.setdefault(tid, Tutor(id=tid, name=tname))
                if atype in (ActivityType.LECTURE, ActivityType.OTHER) and _n_sub == 1:
                    gid = f"{code}/All"
                    if gid not in seen_groups:
                        groups.append(Group(id=gid, course_code=code, label="All", size=size))
                        seen_groups.add(gid)
                    course.activities.append(Activity(
                        course_code=code, activity_type=atype, delivery_mode=mode,
                        duration_slots=duration, weeks=wk,
                        tutor_id=tid, group_id=gid, size=size,
                        fixed_day=fday, fixed_start_index=fstart,
                        is_evening=is_evening,
                        notes=str(slot["remarks"] or ""), co_tutor_ids=[],
                        weeks_from_default=slot.get("_has_empty_weeks", False) and not slot.get("_has_explicit_weeks", False)))
                else:
                    prefix = ("G" if atype in (ActivityType.LECTURE, ActivityType.OTHER)
                              else _group_prefix(atype))
                    for k, (label, gsize) in enumerate(_split_into_n_groups(size, _n_sub, prefix)):
                        gid = f"{code}/{label}"
                        if gid not in seen_groups:
                            groups.append(Group(id=gid, course_code=code, label=label, size=gsize))
                            seen_groups.add(gid)
                        pin_day = fday if label.endswith("1") else None
                        pin_start = fstart if label.endswith("1") else None
                        course.activities.append(Activity(
                            course_code=code, activity_type=atype, delivery_mode=mode,
                            duration_slots=duration, weeks=wk,
                            tutor_id=tid, group_id=gid, size=gsize,
                            fixed_day=pin_day, fixed_start_index=pin_start,
                            is_evening=is_evening,
                            notes=str(slot["remarks"] or ""), co_tutor_ids=[],
                            weeks_from_default=slot.get("_has_empty_weeks", False) and not slot.get("_has_explicit_weeks", False)))

    # ---- rooms + calendar -----------------------------------------------
    # Rooms MUST live in the worksheet's own 'Rooms' tab.  We refuse to
    # silently fall back to inputs.xlsx or template 2.xlsx — every run prints
    # exactly which file the rooms came from so the planner can verify.
    import sys as _sys
    if "Rooms" in wb.sheetnames:
        rooms = _read_inputs_rooms(worksheet_path)
        print(f"[loader] rooms loaded from worksheet: {worksheet_path} "
              f"(Rooms tab, {len(rooms)} rooms)", file=_sys.stderr)
    else:
        raise SystemExit(
            f"{worksheet_path} has no 'Rooms' tab. "
            f"Rooms must live in the same workbook as the modules — "
            f"template 2.xlsx is for syncing OUTPUT only and is never read "
            f"for room data. Add a 'Rooms' tab to {worksheet_path}.")

    # Calendar: prefer a Calendar tab in THIS workbook.
    if "Calendar" in wb.sheetnames:
        cal_kv = {row[0]: row[1] for row in wb["Calendar"].iter_rows(values_only=True) if row and row[0]}
        raw_start = cal_kv.get("Semester start date")
        if raw_start is not None and hasattr(raw_start, "year"):
            sem_start = raw_start if isinstance(raw_start, date) else date.fromisoformat(str(raw_start)[:10])
        elif raw_start:
            sem_start = date.fromisoformat(str(raw_start)[:10])
        else:
            sem_start = date.fromisoformat(str(semester_start)[:10])
        weeks_list = _parse_weeks_range(cal_kv.get("Teaching weeks") or teaching_weeks)
    else:
        sem_start = date.fromisoformat(str(semester_start)[:10])
        weeks_list = _parse_weeks_range(teaching_weeks)
    week_dates = {w: (sem_start + timedelta(days=(w - 1) * 7)).isoformat() for w in weeks_list}
    calendar = Calendar(teaching_weeks=weeks_list, week_dates=week_dates)

    if mode_warnings:
        import sys as _sys
        print("[validation] activity/delivery-mode rule violations:", file=_sys.stderr)
        for _w in mode_warnings:
            print("  - " + _w, file=_sys.stderr)
    if staff_warnings:
        import sys as _sys
        for _w in staff_warnings:
            print("  - " + _w, file=_sys.stderr)

    courses = sorted(by_code.values(), key=lambda c: (c.year, c.code))

    # If any activity is marked as evening (column M ticked), the Settings tab's
    # "Day end hour" may be 18:00 which is too early for a 19:00 start.
    # Extend the day to 22:00 automatically so the slot grid fits evening classes.
    global DAY_END_HOUR
    has_evening = any(a.is_evening for c in courses for a in c.activities)
    if has_evening and DAY_END_HOUR < 22:
        DAY_END_HOUR = 22

    return Universe(
        courses=courses, rooms=rooms,
        tutors=sorted(tutors_by_id.values(), key=lambda t: t.name),
        groups=groups, time_slots=_build_time_slots(), calendar=calendar,
    )


# --- end of data_loader.py ---
